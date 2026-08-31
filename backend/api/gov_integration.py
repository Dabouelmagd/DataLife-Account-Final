"""
Government Integration Engine — محرك التكامل مع المنظومات الحكومية المصرية

1. منظومة البيرول الموحدة — Unified Payroll System (مصلحة الضرائب)
   تصدير JSON/Excel بالصيغة المعتمدة للرفع المباشر

2. نافذة ACI / الجمركية — Advanced Cargo Information
   ربط ACID Number بطلبات الشراء وفواتير الاستيراد

3. المدفوعات الحكومية الإلكترونية — GPS (قانون 18/2019)
   دعم بوابات الدفع الحكومية والشيكات الإلكترونية
"""
import uuid, io
from datetime import datetime, timezone, date
from typing import Optional, List
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel, Field

from database import db
from api.users import get_current_user

router = APIRouter(prefix="/api/gov", tags=["Government Integration"])


# ══════════════════════════════════════════════════════════════
# 1. UNIFIED PAYROLL SYSTEM — منظومة البيرول الموحدة
#    مصلحة الضرائب المصرية — تصدير الرواتب للرفع الإلكتروني
# ══════════════════════════════════════════════════════════════

@router.get("/payroll/unified-export/{run_id}")
async def export_unified_payroll_json(
    run_id: str,
    current_user: dict = Depends(get_current_user)
):
    """
    تصدير كشف الرواتب بصيغة JSON المعتمدة من مصلحة الضرائب
    للرفع المباشر على منظومة البيرول الموحدة
    """
    company_id = current_user["company_id"]

    run = await db.payroll_runs.find_one(
        {"id": run_id, "company_id": company_id}, {"_id": 0})
    if not run:
        raise HTTPException(404, "كشف الرواتب غير موجود")

    company = await db.companies.find_one({"id": company_id}, {"_id": 0}) or {}

    employees_data = run.get("employees_data", [])
    employees_payload = []

    for emp in employees_data:
        # Fetch insurance profile for tax_id and insurance_number
        profile = await db.employee_insurance_profiles.find_one(
            {"employee_id": emp.get("employee_id",""),
             "company_id": company_id}, {"_id": 0}
        ) or {}

        employees_payload.append({
            # ── بيانات هوية الموظف ───────────────────────
            "nationalId":          profile.get("national_id", emp.get("national_id","")),
            "employeeId":          emp.get("employee_id",""),
            "employeeName":        emp.get("employee_name",""),
            "insuranceNumber":     profile.get("insurance_number",""),
            "bankIBAN":            profile.get("bank_account_iban",""),
            # ── بيانات الأجر ────────────────────────────
            "basicSalary":         float(emp.get("basic_salary", 0)),
            "variableElements":    float(emp.get("allowances", 0)),
            "grossSalary":         float(emp.get("gross_salary", 0)),
            # ── التأمينات الاجتماعية ─────────────────────
            "insurableBasic":      float(profile.get("insured_basic_salary", emp.get("basic_salary",0))),
            "insurableVariable":   float(profile.get("insured_variable_salary", 0)),
            "employeeSI":          float(emp.get("employee_si", 0)),
            "employerSI":          float(emp.get("employer_si", 0)),
            # ── ضريبة كسب العمل ─────────────────────────
            "annualTaxableIncome": float(emp.get("annual_taxable", 0)),
            "personalExemption":   float(emp.get("personal_exemption", 21000)),
            "incomeTax":           float(emp.get("income_tax", 0)),
            # ── الصرف الفعلي ────────────────────────────
            "deductions":          float(emp.get("total_deductions", 0)),
            "netSalary":           float(emp.get("net_salary", 0)),
            # ── الخصومات التفصيلية ───────────────────────
            "loanDeduction":       float(emp.get("loan_deduction", 0)),
            "otherDeductions":     float(emp.get("other_deductions", 0)),
        })

    payload = {
        # ── بيانات جهة العمل ──────────────────────────────
        "submissionType":    "monthly",
        "submissionVersion": "2.0",
        "taxAuthority":      "ETA",
        "employer": {
            "taxRegistrationNumber": company.get("tax_id",""),
            "companyName":           company.get("name",""),
            "companyNameEn":         company.get("name_en",""),
            "insuranceRegNumber":    company.get("insurance_registration_number",""),
            "activityCode":          company.get("activity_code",""),
            "address":               company.get("address",""),
        },
        # ── بيانات الفترة ─────────────────────────────────
        "period": {
            "year":      run.get("year", date.today().year),
            "month":     run.get("month", date.today().month),
            "periodType":"monthly",
        },
        # ── إجماليات ─────────────────────────────────────
        "totals": {
            "employeeCount":   len(employees_payload),
            "totalGross":      round(sum(e["grossSalary"] for e in employees_payload), 2),
            "totalEmployeeSI": round(sum(e["employeeSI"] for e in employees_payload), 2),
            "totalEmployerSI": round(sum(e["employerSI"] for e in employees_payload), 2),
            "totalIncomeTax":  round(sum(e["incomeTax"] for e in employees_payload), 2),
            "totalNetPaid":    round(sum(e["netSalary"] for e in employees_payload), 2),
        },
        # ── تفصيل الموظفين ───────────────────────────────
        "employees": employees_payload,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "generatedBy":  current_user["user_id"],
    }

    return JSONResponse(
        content=payload,
        headers={
            "Content-Disposition": f'attachment; filename="payroll_{run.get("year")}_{run.get("month"):02d}.json"',
            "X-Payroll-Run-Id":    run_id,
            "X-Employee-Count":    str(len(employees_payload)),
        }
    )


@router.get("/payroll/unified-export/{run_id}/excel")
async def export_unified_payroll_excel(
    run_id: str,
    current_user: dict = Depends(get_current_user)
):
    """
    تصدير كشف الرواتب بصيغة Excel المعتمدة من مصلحة الضرائب
    النموذج القياسي لرفع ملف البيرول الشهري
    """
    company_id = current_user["company_id"]

    run = await db.payroll_runs.find_one(
        {"id": run_id, "company_id": company_id}, {"_id": 0})
    if not run:
        raise HTTPException(404, "كشف الرواتب غير موجود")

    company = await db.companies.find_one({"id": company_id}, {"_id": 0}) or {}
    employees_data = run.get("employees_data", [])

    # Build Excel using openpyxl
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "مكتبة openpyxl غير مثبتة — pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"بيرول {run.get('year')}-{run.get('month',0):02d}"

    # ── Company header ─────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    sub_fill    = PatternFill("solid", fgColor="2E86AB")

    ws.merge_cells("A1:R1")
    ws["A1"] = f"منظومة البيرول الموحدة — مصلحة الضرائب المصرية"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:R2")
    ws["A2"] = (f"{company.get('name','')} | رقم ضريبي: {company.get('tax_id','')} | "
                f"الفترة: {run.get('year')}/{run.get('month',0):02d}")
    ws["A2"].fill = sub_fill
    ws["A2"].font = Font(color="FFFFFF", bold=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── Column headers ─────────────────────────────────────
    col_fill = PatternFill("solid", fgColor="F0F4F8")
    col_font = Font(bold=True, size=10)
    border   = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    headers = [
        "الرقم القومي", "رقم الموظف", "اسم الموظف",
        "رقم التأمين", "IBAN",
        "الراتب الأساسي", "البدلات والمتغيرات", "إجمالي الراتب",
        "وعاء التأمين (أساسي)", "وعاء التأمين (متغير)",
        "تأمين الموظف", "تأمين صاحب العمل",
        "الوعاء الضريبي السنوي", "الإعفاء الشخصي", "ضريبة كسب العمل",
        "إجمالي الخصومات", "صافي الراتب المصروف",
        "ملاحظات",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = col_fill
        cell.font = col_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    ws.row_dimensions[3].height = 35

    # ── Employee rows ──────────────────────────────────────
    alt_fill = PatternFill("solid", fgColor="F7FAFC")
    num_fmt  = '#,##0.00'

    for row_idx, emp in enumerate(employees_data, 4):
        profile = await db.employee_insurance_profiles.find_one(
            {"employee_id": emp.get("employee_id",""), "company_id": company_id},
            {"_id": 0}
        ) or {}

        row_fill = alt_fill if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        row_data = [
            profile.get("national_id", emp.get("national_id","")),
            emp.get("employee_id",""),
            emp.get("employee_name",""),
            profile.get("insurance_number",""),
            profile.get("bank_account_iban",""),
            float(emp.get("basic_salary",0)),
            float(emp.get("allowances",0)),
            float(emp.get("gross_salary",0)),
            float(profile.get("insured_basic_salary", emp.get("basic_salary",0))),
            float(profile.get("insured_variable_salary",0)),
            float(emp.get("employee_si",0)),
            float(emp.get("employer_si",0)),
            float(emp.get("annual_taxable",0)),
            float(emp.get("personal_exemption",21000)),
            float(emp.get("income_tax",0)),
            float(emp.get("total_deductions",0)),
            float(emp.get("net_salary",0)),
            "",
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.fill   = row_fill
            if isinstance(value, float):
                cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal="right")

    # ── Totals row ─────────────────────────────────────────
    total_row = len(employees_data) + 4
    ws.merge_cells(f"A{total_row}:E{total_row}")
    ws[f"A{total_row}"] = f"الإجماليات ({len(employees_data)} موظف)"
    ws[f"A{total_row}"].font = Font(bold=True)
    ws[f"A{total_row}"].fill = PatternFill("solid", fgColor="E8F4F8")

    total_cols = {6: "basic_salary", 7: "allowances", 8: "gross_salary",
                  11: "employee_si", 12: "employer_si", 15: "income_tax",
                  16: "total_deductions", 17: "net_salary"}
    for col, field in total_cols.items():
        total_val = round(sum(float(e.get(field,0)) for e in employees_data), 2)
        cell = ws.cell(row=total_row, column=col, value=total_val)
        cell.font = Font(bold=True)
        cell.number_format = num_fmt
        cell.fill = PatternFill("solid", fgColor="E8F4F8")

    # ── Save and stream ────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"payroll_unified_{run.get('year')}_{run.get('month',0):02d}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/payroll/quarterly-form4/{year}/{quarter}")
async def export_quarterly_form4(
    year: int, quarter: int,
    current_user: dict = Depends(get_current_user)
):
    """
    تصدير نموذج 4 الربع سنوي — مصلحة الضرائب
    صيغة JSON للرفع الإلكتروني على بوابة الضرائب
    """
    if quarter not in (1,2,3,4):
        raise HTTPException(400, "الربع يجب أن يكون 1-4")

    company_id = current_user["company_id"]
    company    = await db.companies.find_one({"id": company_id}, {"_id": 0}) or {}

    month_start = (quarter - 1) * 3 + 1
    month_end   = quarter * 3

    runs = await db.payroll_runs.find({
        "company_id": company_id,
        "year":  year,
        "month": {"$gte": month_start, "$lte": month_end},
        "status": {"$in": ["approved","paid"]},
    }, {"_id": 0}).to_list(None)

    emp_totals = {}
    for run in runs:
        for emp in run.get("employees_data", []):
            eid = emp.get("employee_id","")
            if eid not in emp_totals:
                profile = await db.employee_insurance_profiles.find_one(
                    {"employee_id": eid, "company_id": company_id}, {"_id": 0}) or {}
                emp_totals[eid] = {
                    "nationalId":      profile.get("national_id",""),
                    "employeeName":    emp.get("employee_name",""),
                    "insuranceNumber": profile.get("insurance_number",""),
                    "grossSalary": 0.0, "incomeTax": 0.0,
                    "employeeSI": 0.0,  "netSalary": 0.0,
                }
            emp_totals[eid]["grossSalary"] += float(emp.get("gross_salary",0))
            emp_totals[eid]["incomeTax"]   += float(emp.get("income_tax",0))
            emp_totals[eid]["employeeSI"]  += float(emp.get("employee_si",0))
            emp_totals[eid]["netSalary"]   += float(emp.get("net_salary",0))

    employees = list(emp_totals.values())
    total_tax  = round(sum(e["incomeTax"] for e in employees), 2)
    total_gross= round(sum(e["grossSalary"] for e in employees), 2)

    return {
        "formType":      "Form4-QuarterlyPayrollTax",
        "taxAuthority":  "ETA",
        "version":       "2.0",
        "employer": {
            "taxRegistrationNumber": company.get("tax_id",""),
            "companyName":           company.get("name",""),
            "insuranceRegNumber":    company.get("insurance_registration_number",""),
        },
        "period": {
            "year":    year,
            "quarter": quarter,
            "months":  list(range(month_start, month_end+1)),
        },
        "employees":    employees,
        "totals": {
            "employeeCount":  len(employees),
            "totalGross":     total_gross,
            "totalIncomeTax": total_tax,
        },
        "submissionNote": f"يُرفَع على بوابة مصلحة الضرائب خلال 30 يوماً من نهاية الربع",
        "generatedAt":    datetime.now(timezone.utc).isoformat(),
    }


# ══════════════════════════════════════════════════════════════
# 2. ACI — نافذة التسجيل المسبق للشحنات (ACI / الجمركية)
#    Advance Cargo Information — ACID Number Integration
# ══════════════════════════════════════════════════════════════

class ACIRegistrationRequest(BaseModel):
    document_id:       str          # ID of PO or import invoice
    document_type:     str = "purchase_order"  # purchase_order | import_invoice
    acid_number:       str          # رقم ACID من نافذة الجمركية
    shipment_date:     str
    vessel_name:       Optional[str] = None
    port_of_loading:   Optional[str] = None
    port_of_discharge: Optional[str] = "Alexandria"
    hs_codes:          List[str] = []   # Harmonized System codes
    total_weight_kg:   Optional[float] = None
    total_value_usd:   Optional[float] = None
    customs_declaration_number: Optional[str] = None
    notes:             Optional[str] = None


@router.post("/aci/register")
async def register_aci(
    req: ACIRegistrationRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    تسجيل رقم ACID لشحنة مستوردة وربطه بطلب الشراء أو فاتورة الاستيراد

    يُحفَظ رقم ACID مع بيانات الشحنة لاستخدامه في:
    - الجمارك عند وصول البضاعة
    - تسوية فواتير الاستيراد
    - Landed costs distribution
    """
    company_id = current_user["company_id"]

    # Validate ACID format (Egyptian customs: typically 16 digits)
    acid = req.acid_number.strip().replace("-", "").replace(" ", "")
    if len(acid) < 8:
        raise HTTPException(400, f"رقم ACID غير صحيح: '{req.acid_number}' — يجب أن يكون 8+ أرقام")

    # Check document exists
    if req.document_type == "purchase_order":
        doc = await db.purchase_orders.find_one(
            {"id": req.document_id, "company_id": company_id}, {"_id": 0})
    else:
        doc = await db.invoices.find_one(
            {"id": req.document_id, "company_id": company_id}, {"_id": 0})

    if not doc:
        raise HTTPException(404, f"المستند {req.document_id} غير موجود")

    aci_id  = str(uuid.uuid4())
    aci_rec = {
        "id":            aci_id,
        "company_id":    company_id,
        "document_id":   req.document_id,
        "document_type": req.document_type,
        "acid_number":   acid,
        "shipment_date": req.shipment_date,
        "vessel_name":   req.vessel_name,
        "port_of_loading":   req.port_of_loading,
        "port_of_discharge": req.port_of_discharge,
        "hs_codes":      req.hs_codes,
        "total_weight_kg":   req.total_weight_kg,
        "total_value_usd":   req.total_value_usd,
        "customs_declaration_number": req.customs_declaration_number,
        "status":        "registered",  # registered | cleared | rejected
        "notes":         req.notes,
        "registered_by": current_user["user_id"],
        "registered_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.aci_registrations.insert_one(aci_rec)
    aci_rec.pop("_id", None)

    # Update the source document with ACID number
    collection = db.purchase_orders if req.document_type == "purchase_order" else db.invoices
    await collection.update_one(
        {"id": req.document_id},
        {"$set": {
            "acid_number":   acid,
            "aci_id":        aci_id,
            "aci_registered": True,
            "shipment_date": req.shipment_date,
        }}
    )

    return {
        "message":    f"✅ تم تسجيل رقم ACID على {req.document_type}",
        "aci_id":     aci_id,
        "acid_number": acid,
        "document_id": req.document_id,
        "aci_record": aci_rec,
        "customs_note": "يجب تقديم رقم ACID لمصلحة الجمارك قبل وصول الشحنة بـ 48 ساعة",
    }


@router.put("/aci/{aci_id}/clear")
async def clear_aci_shipment(
    aci_id: str, data: dict,
    current_user: dict = Depends(get_current_user)
):
    """تحديث حالة الشحنة بعد التخليص الجمركي"""
    company_id = current_user["company_id"]
    customs_decl = data.get("customs_declaration_number","")
    cleared_date = data.get("cleared_date", date.today().isoformat())

    result = await db.aci_registrations.update_one(
        {"id": aci_id, "company_id": company_id},
        {"$set": {
            "status": "cleared",
            "customs_declaration_number": customs_decl,
            "cleared_date": cleared_date,
            "cleared_by":   current_user["user_id"],
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(404, "سجل ACI غير موجود")

    return {
        "message":     "✅ تم تسجيل التخليص الجمركي",
        "aci_id":      aci_id,
        "customs_decl": customs_decl,
        "cleared_date": cleared_date,
    }


@router.get("/aci/pending")
async def list_pending_aci(current_user: dict = Depends(get_current_user)):
    """قائمة الشحنات المسجلة وغير المُخلَّصة جمركياً"""
    pending = await db.aci_registrations.find({
        "company_id": current_user["company_id"],
        "status": "registered",
    }, {"_id": 0}).sort("shipment_date", 1).to_list(None)
    return {
        "pending_shipments": pending,
        "count":             len(pending),
        "note":              "هذه الشحنات تحتاج تخليصاً جمركياً",
    }


@router.get("/aci/document/{document_id}")
async def get_aci_by_document(
    document_id: str,
    current_user: dict = Depends(get_current_user)
):
    """جلب بيانات ACI المرتبطة بمستند شراء أو فاتورة استيراد"""
    aci = await db.aci_registrations.find_one(
        {"document_id": document_id, "company_id": current_user["company_id"]},
        {"_id": 0}
    )
    if not aci:
        return {"exists": False, "document_id": document_id,
                "message": "لا يوجد رقم ACID مسجل — استخدم POST /aci/register"}
    return {"exists": True, "aci": aci}


# ══════════════════════════════════════════════════════════════
# 3. GOVERNMENT PAYMENT SYSTEM — نظام المدفوعات الحكومية GPS
#    قانون رقم 18 لسنة 2019 — الدفع الإلكتروني الإلزامي
# ══════════════════════════════════════════════════════════════

GPS_GATEWAYS = {
    "tax_authority":    "مصلحة الضرائب المصرية (ETA)",
    "social_insurance": "الهيئة القومية للتأمين الاجتماعي (NOSI)",
    "customs":          "مصلحة الجمارك",
    "real_estate_tax":  "مصلحة الضرائب العقارية",
    "ministry_finance": "وزارة المالية",
    "local_units":      "الوحدات المحلية",
}

GPS_PAYMENT_TYPES = {
    "income_tax":       {"gateway": "tax_authority",    "form": "نموذج 41"},
    "vat":              {"gateway": "tax_authority",    "form": "نموذج 10"},
    "payroll_tax":      {"gateway": "tax_authority",    "form": "نموذج 4"},
    "social_insurance": {"gateway": "social_insurance", "form": "تسوية شهرية"},
    "customs_duty":     {"gateway": "customs",          "form": "بيان جمركي"},
    "real_estate_tax":  {"gateway": "real_estate_tax",  "form": "إقرار ضريبة عقارية"},
    "stamp_duty":       {"gateway": "tax_authority",    "form": "دمغة"},
}


class GovernmentPaymentRequest(BaseModel):
    payment_type:     str   # income_tax | vat | payroll_tax | social_insurance | customs_duty
    amount:           float
    period:           str   # "2026-01" or "2026-Q1"
    reference_number: Optional[str] = None   # رقم المرجع من البوابة
    payment_date:     str
    bank_account:     str = "112"
    notes:            Optional[str] = None


@router.post("/gps/payment")
async def record_government_payment(
    req: GovernmentPaymentRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    تسجيل دفعة حكومية وفق قانون 18/2019

    يُنشئ القيد المحاسبي:
    Dr حساب الضريبة / التأمين المستحق
    Cr البنك

    ويحفظ سجل الدفع مع رقم المرجع الحكومي
    """
    company_id = current_user["company_id"]
    payment_id = str(uuid.uuid4())

    if req.payment_type not in GPS_PAYMENT_TYPES:
        raise HTTPException(400,
            f"نوع الدفع غير مدعوم. الأنواع: {list(GPS_PAYMENT_TYPES.keys())}")

    config  = GPS_PAYMENT_TYPES[req.payment_type]
    gateway = GPS_GATEWAYS.get(config["gateway"], config["gateway"])

    # Map payment type to liability account
    liability_accounts = {
        "income_tax":       ("254", "ضريبة دخل مستحقة"),
        "vat":              ("260", "VAT مخرجات مستحقة"),
        "payroll_tax":      ("261", "ضريبة كسب عمل مستحقة"),
        "social_insurance": ("255", "تأمينات اجتماعية مستحقة"),
        "customs_duty":     ("254", "رسوم جمركية مستحقة"),
        "real_estate_tax":  ("2541","ضريبة تصرفات عقارية"),
        "stamp_duty":       ("254", "دمغة مستحقة"),
    }
    dr_code, dr_name = liability_accounts.get(req.payment_type, ("254","ضريبة مستحقة"))

    # Journal entry
    from services.accounting_service import AccountingService
    from models.accounting import JournalEntry
    svc = AccountingService(db)

    async def get_acc(code):
        a = await db.chart_of_accounts.find_one(
            {"company_id": company_id, "account_code": code}, {"_id": 0})
        return a or {"id": code, "account_code": code, "account_name": f"حساب {code}"}

    dr_acc = await get_acc(dr_code)
    cr_acc = await get_acc(req.bank_account)

    lines = [
        {
            "line_id": str(uuid.uuid4()), "entry_id": None,
            "account_id": dr_acc["id"], "account_code": dr_code,
            "account_name": dr_name,
            "debit": req.amount, "credit": 0,
            "description": f"سداد {dr_name} — {req.period} — {gateway}",
        },
        {
            "line_id": str(uuid.uuid4()), "entry_id": None,
            "account_id": cr_acc["id"], "account_code": req.bank_account,
            "account_name": "البنك",
            "debit": 0, "credit": req.amount,
            "description": f"تحويل بنكي لـ {gateway} — {req.period}",
        },
    ]

    entry = JournalEntry(
        company_id=company_id, entry_number=0,
        entry_date=req.payment_date,
        description=f"دفعة حكومية GPS — {dr_name} — {req.period}",
        lines=lines, source_document_type="manual",
        source_document_id=payment_id,
        created_by=current_user["user_id"],
    )
    je_result = await svc.create_journal_entry(entry)
    await svc.post_journal_entry(je_result["id"], current_user["user_id"])

    # Save payment record
    payment = {
        "id":                payment_id,
        "company_id":        company_id,
        "payment_type":      req.payment_type,
        "payment_type_ar":   dr_name,
        "gateway":           gateway,
        "form_reference":    config.get("form",""),
        "amount":            req.amount,
        "period":            req.period,
        "payment_date":      req.payment_date,
        "reference_number":  req.reference_number,
        "bank_account":      req.bank_account,
        "journal_entry_id":  je_result["id"],
        "law_reference":     "قانون رقم 18 لسنة 2019 — الدفع الإلكتروني الإلزامي",
        "notes":             req.notes,
        "recorded_by":       current_user["user_id"],
        "recorded_at":       datetime.now(timezone.utc).isoformat(),
    }
    await db.government_payments.insert_one(payment)
    payment.pop("_id", None)

    return {
        "message":        f"✅ تم تسجيل الدفعة الحكومية — {gateway}",
        "payment_id":     payment_id,
        "payment":        payment,
        "journal": {
            "id":     je_result["id"],
            "debit":  f"م/{dr_code} {dr_name}  {req.amount:,.2f}",
            "credit": f"م/{req.bank_account} البنك  {req.amount:,.2f}",
        },
    }


@router.post("/gps/electronic-cheque")
async def issue_electronic_cheque(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    إصدار شيك إلكتروني للجهات الحكومية
    وفق اشتراطات قانون 18/2019 ونظام GPS

    يُولِّد بيانات الشيك الإلكتروني بالصيغة المطلوبة
    """
    company_id  = current_user["company_id"]
    company     = await db.companies.find_one({"id": company_id}, {"_id": 0}) or {}
    cheque_id   = str(uuid.uuid4())
    cheque_num  = f"EGOV-{date.today().year}-{str(uuid.uuid4())[:6].upper()}"

    amount       = float(data.get("amount", 0))
    beneficiary  = data.get("beneficiary","")  # الجهة الحكومية
    payment_type = data.get("payment_type","")
    period       = data.get("period","")
    issue_date   = data.get("issue_date", date.today().isoformat())

    cheque_payload = {
        # ── بيانات المُصدِر ────────────────────────────────
        "issuerData": {
            "taxRegistrationNumber": company.get("tax_id",""),
            "companyName":           company.get("name",""),
            "bankAccount":           data.get("bank_account",""),
            "bankName":              data.get("bank_name","البنك الأهلي المصري"),
        },
        # ── بيانات الشيك ──────────────────────────────────
        "chequeData": {
            "chequeNumber":  cheque_num,
            "issueDate":     issue_date,
            "amount":        amount,
            "amountInWords": _amount_to_arabic_words(amount),
            "currency":      "EGP",
            "beneficiary":   beneficiary,
            "paymentType":   payment_type,
            "period":        period,
            "memo":          data.get("memo",""),
        },
        # ── بيانات التحقق ─────────────────────────────────
        "verificationData": {
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "generatedBy": current_user["user_id"],
            "systemRef":   cheque_id,
            "lawReference":"قانون 18/2019 — الشيكات الإلكترونية الحكومية",
        }
    }

    record = {
        "id":             cheque_id,
        "company_id":     company_id,
        "cheque_number":  cheque_num,
        "amount":         amount,
        "beneficiary":    beneficiary,
        "payment_type":   payment_type,
        "period":         period,
        "issue_date":     issue_date,
        "status":         "issued",
        "payload":        cheque_payload,
        "created_by":     current_user["user_id"],
        "created_at":     datetime.now(timezone.utc).isoformat(),
    }
    await db.electronic_cheques.insert_one(record)
    record.pop("_id", None)
    record.pop("payload", None)

    return {
        "message":       f"✅ تم إصدار شيك إلكتروني لـ {beneficiary}",
        "cheque_id":     cheque_id,
        "cheque_number": cheque_num,
        "amount":        amount,
        "cheque_payload": cheque_payload,
    }


@router.get("/gps/payments")
async def list_government_payments(
    payment_type: Optional[str] = None,
    year:         Optional[int] = None,
    current_user: dict = Depends(get_current_user)
):
    """سجل الدفعات الحكومية"""
    q = {"company_id": current_user["company_id"]}
    if payment_type: q["payment_type"] = payment_type
    if year:         q["payment_date"] = {"$gte": f"{year}-01-01", "$lte": f"{year}-12-31"}

    payments = await db.government_payments.find(q, {"_id": 0}).sort(
        "payment_date", -1).to_list(None)
    total = round(sum(float(p.get("amount",0)) for p in payments), 2)

    return {
        "payments":   payments,
        "count":      len(payments),
        "total_paid": total,
        "gateways":   GPS_GATEWAYS,
    }


@router.get("/gps/gateways")
async def list_gps_gateways(current_user: dict = Depends(get_current_user)):
    """قائمة بوابات الدفع الحكومية المدعومة"""
    return {
        "gateways":      GPS_GATEWAYS,
        "payment_types": GPS_PAYMENT_TYPES,
        "law_reference": "قانون رقم 18 لسنة 2019 — إلزامية الدفع الإلكتروني للجهات الحكومية",
        "note":          "جميع الدفعات للجهات الحكومية > 500 ج.م يجب أن تكون إلكترونية",
    }


def _amount_to_arabic_words(amount: float) -> str:
    """تحويل المبلغ الرقمي إلى كلمات عربية (مبسط)"""
    int_part  = int(amount)
    frac_part = round((amount - int_part) * 100)
    result    = f"فقط {int_part:,} جنيه مصري"
    if frac_part > 0:
        result += f" و{frac_part} قرش"
    result += " لا غير"
    return result
